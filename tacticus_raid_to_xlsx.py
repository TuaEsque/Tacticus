import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

FONT = "Arial"
DARK_BLUE = "1F4E79"
MID_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ALT_ROW = "EBF3FB"
WHITE = "FFFFFF"
GOLD = "FFD700"
SILVER = "C0C0C0"
BRONZE = "CD7F32"


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, value, bg=MID_BLUE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    return c


def _cell(ws, row, col, value, bg=WHITE, bold=False, fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, bold=bold, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c


def _title(ws, text, num_cols):
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name=FONT, bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _row_bg(i):
    return ALT_ROW if i % 2 == 1 else WHITE


# ── JSON Parsing ───────────────────────────────────────────────────────────────

def parse_json(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    users, entries = {}, []

    for uid, udata in data.get("users", {}).items():
        rarities = {}
        for rname, rdata in udata.get("rarities", {}).items():
            bosses = {}
            for bname, bdata in rdata.get("bosses", {}).items():
                dtypes = {}
                for dname, dtdata in bdata.get("damageTypes", {}).items():
                    dt_entries = []
                    for entry in dtdata.get("entries", []):
                        e = {
                            "userId": uid,
                            "boss": bname,
                            "rarity": rname,
                            "damageType": dname,
                            "damageDealt": int(entry.get("damageDealt") or 0),
                            "encounterType": entry.get("encounterType"),
                            "unitId": entry.get("unitId"),
                            "type": entry.get("type"),
                            "startedOn": entry.get("startedOn"),
                            "completedOn": entry.get("completedOn"),
                            "team": entry.get("team"),
                        }
                        dt_entries.append(e)
                        entries.append(e)
                    dtypes[dname] = {
                        "total": int(dtdata.get("totalDamage") or 0),
                        "entries": dt_entries,
                    }
                bosses[bname] = {
                    "total": int(bdata.get("totalDamage") or 0),
                    "damageTypes": dtypes,
                }
            rarities[rname] = {
                "total": int(rdata.get("totalDamage") or 0),
                "bosses": bosses,
            }
        users[uid] = {
            "total": int(udata.get("totalDamage") or 0),
            "rarities": rarities,
        }

    return users, entries


# ── Sheet 1: Summary ───────────────────────────────────────────────────────────

def _summary(wb, users):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    _title(ws, "Guild Raid – Player Summary", 4)

    for col, label in enumerate(["Rank", "Player ID", "Total Damage", "% of Guild"], 1):
        _hdr(ws, 2, col, label)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    by_damage = sorted(users.items(), key=lambda x: x[1]["total"], reverse=True)
    tr = len(by_damage) + 3
    rank_bg = {1: GOLD, 2: SILVER, 3: BRONZE}

    for i, (uid, udata) in enumerate(by_damage, 3):
        rank = i - 2
        bg = rank_bg.get(rank, _row_bg(i))
        _cell(ws, i, 1, rank, bg=bg, align="center")
        _cell(ws, i, 2, uid, bg=bg)
        _cell(ws, i, 3, udata["total"], bg=bg, fmt="#,##0", align="right")
        _cell(ws, i, 4, f"=IFERROR(C{i}/C${tr},0)", bg=bg, fmt="0.0%", align="right")

    ws.merge_cells(f"A{tr}:B{tr}")
    _cell(ws, tr, 1, "GUILD TOTAL", bg=LIGHT_BLUE, bold=True, align="right")
    _cell(ws, tr, 3, f"=SUM(C3:C{tr - 1})", bg=LIGHT_BLUE, bold=True, fmt="#,##0", align="right")
    _cell(ws, tr, 4, 1.0, bg=LIGHT_BLUE, bold=True, fmt="0.0%", align="right")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14


# ── Sheet 2: User vs Boss matrix ───────────────────────────────────────────────

def _boss_matrix(wb, users):
    ws = wb.create_sheet("User vs Boss")
    ws.sheet_view.showGridLines = False

    bosses = sorted({b for u in users.values() for r in u["rarities"].values() for b in r["bosses"]})
    by_damage = sorted(users.items(), key=lambda x: x[1]["total"], reverse=True)
    tcol = len(bosses) + 2

    _title(ws, "Guild Raid – Damage by Player & Boss", tcol)
    _hdr(ws, 2, 1, "Player ID", bg=DARK_BLUE)
    for i, b in enumerate(bosses, 2):
        _hdr(ws, 2, i, b)
    _hdr(ws, 2, tcol, "Total", bg=DARK_BLUE)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for i, (uid, udata) in enumerate(by_damage, 3):
        bg = _row_bg(i)
        _cell(ws, i, 1, uid, bg=bg)
        fc = get_column_letter(2)
        lc = get_column_letter(tcol - 1)
        for j, b in enumerate(bosses, 2):
            val = sum(r["bosses"].get(b, {}).get("total", 0) for r in udata["rarities"].values())
            _cell(ws, i, j, val if val else None, bg=bg, fmt="#,##0", align="right")
        _cell(ws, i, tcol, f"=SUM({fc}{i}:{lc}{i})",
              bg=LIGHT_BLUE, bold=True, fmt="#,##0", align="right")

    gr = len(by_damage) + 3
    _cell(ws, gr, 1, "TOTAL", bg=LIGHT_BLUE, bold=True, align="right")
    for col in range(2, tcol + 1):
        cl = get_column_letter(col)
        _cell(ws, gr, col, f"=SUM({cl}3:{cl}{gr - 1})",
              bg=LIGHT_BLUE, bold=True, fmt="#,##0", align="right")

    if by_damage and bosses:
        data_range = f"B3:{get_column_letter(tcol - 1)}{gr - 1}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="BDD7EE",
                end_type="max", end_color=MID_BLUE,
            ),
        )

    ws.column_dimensions["A"].width = 42
    for j, b in enumerate(bosses, 2):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(b) + 2)
    ws.column_dimensions[get_column_letter(tcol)].width = 14


# ── Sheet 3: Damage breakdown ──────────────────────────────────────────────────

def _breakdown(wb, users):
    ws = wb.create_sheet("Damage Breakdown")
    ws.sheet_view.showGridLines = False
    _title(ws, "Guild Raid – Detailed Damage Breakdown", 7)

    for col, label in enumerate(["Player ID", "Rarity", "Boss", "Damage Type",
                                  "Total Damage", "Entries", "% of User Total"], 1):
        _hdr(ws, 2, col, label)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    by_damage = sorted(users.items(), key=lambda x: x[1]["total"], reverse=True)
    row = 3
    for uid, udata in by_damage:
        for rar in sorted(udata["rarities"]):
            for boss in sorted(udata["rarities"][rar]["bosses"]):
                for dt in sorted(udata["rarities"][rar]["bosses"][boss]["damageTypes"]):
                    dtdata = udata["rarities"][rar]["bosses"][boss]["damageTypes"][dt]
                    bg = _row_bg(row)
                    _cell(ws, row, 1, uid, bg=bg)
                    _cell(ws, row, 2, rar, bg=bg)
                    _cell(ws, row, 3, boss, bg=bg)
                    _cell(ws, row, 4, dt, bg=bg)
                    _cell(ws, row, 5, dtdata["total"], bg=bg, fmt="#,##0", align="right")
                    _cell(ws, row, 6, len(dtdata["entries"]), bg=bg, fmt="#,##0", align="right")
                    _cell(
                        ws, row, 7,
                        f"=IFERROR(E{row}/SUMIF($A:$A,$A{row},$E:$E),0)",
                        bg=bg, fmt="0.0%", align="right",
                    )
                    row += 1

    ws.auto_filter.ref = f"A2:G{row - 1}"
    for col, w in zip(range(1, 8), [42, 20, 14, 14, 16, 10, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Sheet 4: Raw entries ───────────────────────────────────────────────────────

def _raw_entries(wb, entries):
    ws = wb.create_sheet("Raw Entries")
    ws.sheet_view.showGridLines = False
    _title(ws, "Guild Raid – All Individual Entries", 10)

    headers = ["Player ID", "Rarity", "Boss", "Damage Type", "Damage Dealt",
               "Encounter Type", "Unit ID", "Type", "Started On", "Completed On"]
    for col, label in enumerate(headers, 1):
        _hdr(ws, 2, col, label)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for i, e in enumerate(entries, 3):
        bg = _row_bg(i)
        vals = [e["userId"], e["rarity"], e["boss"], e["damageType"], e["damageDealt"],
                e["encounterType"], e["unitId"], e["type"], e["startedOn"], e["completedOn"]]
        for col, v in enumerate(vals, 1):
            _cell(ws, i, col, v, bg=bg,
                  fmt="#,##0" if col == 5 else None,
                  align="right" if col == 5 else "left")

    ws.auto_filter.ref = f"A2:J{len(entries) + 2}"
    for col, w in zip(range(1, 11), [42, 20, 14, 14, 14, 14, 18, 14, 20, 20]):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Sheet 5: Battle teams (created only when battle entries exist) ─────────────

def _battle_teams(wb, entries):
    battles = [e for e in entries
               if e.get("team") and e.get("damageType", "").lower() == "battle"]
    if not battles:
        return

    ws = wb.create_sheet("Battle Teams")
    ws.sheet_view.showGridLines = False
    _title(ws, "Guild Raid – Battle Team Compositions", 8)

    for col, label in enumerate(["Player ID", "Rarity", "Boss", "Damage Dealt",
                                  "Heroes", "Hero Powers", "Machine of War", "MoW Power"], 1):
        _hdr(ws, 2, col, label)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for i, e in enumerate(battles, 3):
        bg = _row_bg(i)
        team = e.get("team") or {}
        heroes = team.get("heroes") or []
        machine = team.get("machineOfWar")
        hero_names = ", ".join(h["unitId"] for h in heroes if h.get("unitId"))
        hero_pwrs = ", ".join(str(h["power"]) for h in heroes if h.get("power"))
        mow_id = (machine or {}).get("unitId") or "—"
        mow_pwr = str((machine or {}).get("power") or "")
        for col, v in enumerate([e["userId"], e["rarity"], e["boss"], e["damageDealt"],
                                  hero_names, hero_pwrs, mow_id, mow_pwr], 1):
            _cell(ws, i, col, v, bg=bg,
                  fmt="#,##0" if col == 4 else None,
                  align="right" if col == 4 else "left")

    ws.auto_filter.ref = f"A2:H{len(battles) + 2}"
    for col, w in zip(range(1, 9), [42, 20, 14, 14, 40, 30, 20, 12]):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Convert guild_raid_analysis.json to Excel.")
    parser.add_argument("json_file", nargs="?", default="guild_raid_analysis.json",
                        help="Path to the JSON export (default: guild_raid_analysis.json)")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: same stem as JSON + .xlsx)")
    args = parser.parse_args()

    json_path = Path(args.json_file)
    if not json_path.exists():
        sys.exit(f"Error: {json_path} not found.")

    out_path = Path(args.output) if args.output else json_path.with_suffix(".xlsx")
    print(f"Parsing {json_path} …")
    users, entries = parse_json(json_path)
    if not users:
        sys.exit("No user data found in JSON.")
    print(f"  {len(users)} players, {len(entries)} entries.")

    wb = Workbook()
    wb.remove(wb.active)
    _summary(wb, users)
    _boss_matrix(wb, users)
    _breakdown(wb, users)
    _raw_entries(wb, entries)
    _battle_teams(wb, entries)

    wb.save(out_path)
    print(f"Saved → {out_path}")


if __name__ == "__main__":
    main()